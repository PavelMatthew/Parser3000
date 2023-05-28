from openpyxl import load_workbook
from openpyxl.styles import Font

import time
import requests
from bs4 import BeautifulSoup

class DemoException(Exception):
    pass

#Url = "https://www.houzz.ru/professionaly/dizaynery-interyera/natalyya-preobrazhenskaya-byuro-uyutnaya-kvartira-pfvwru-pf~1704164889"

t_start = time.time()
print(str(t_start))

def ParsingPage(URLneed):

    
        
    response = requests.get(URLneed)
    soup = BeautifulSoup(response.text, "lxml")
    try:
        global BusinessName
        BusinessName = soup.find("div",class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find("p", class_="sc-mwxddt-0 cZJFpr").text
    except:
        BusinessName="Нет имени"
    try:
        global PhoneNumber
        PhoneNumber= soup.find("div",class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find_next_sibling().find("p", class_="sc-mwxddt-0 cZJFpr").text
    except:
        PhoneNumber="Нет номера"
    try:
        global Website
        Website =soup.find("div",class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find_next_sibling().find_next_sibling().find("a", class_="sc-62xgu6-0 cZBXc sc-mwxddt-0 kCqoeY hui-link").text
    except:
        Website="Нет сайта"
    try:
        global Contacts
        Contacts = soup.find("div", class_= "sc-mwxddt-0 hvKGQq").text
    except:
        Contatcs = "Нет контактов"
    try:
        global Address
        if Website =="Нет сайта":
            Address= soup.find("div",class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find_next_sibling().find_next_sibling().find("p",class_="sc-mwxddt-0 cZJFpr").text
        else:
            Address= soup.find("div",class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find_next_sibling().find_next_sibling().find_next_sibling().find("p",class_="sc-mwxddt-0 cZJFpr").text
    except:
        Address = "Нет адреса"
    try:
        global Followers
        
        Followers = soup.find("div",class_="sc-183mtny-0 ePhTrT").find("a",class_="sc-62xgu6-0 cZBXc sc-mwxddt-0 kCqoeY hui-link").text.replace("Подписчик: 1","1").replace("Подписчиков: ","")
        
    except:
        Followers = "0"
        
    try:
        global TypicalJobCost
        if Website =="Нет сайта":
            TypicalJobCost = soup.find("div", class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find_next_sibling().find_next_sibling().find_next_sibling().find("p", class_="sc-mwxddt-0 cZJFpr").text.replace("Подписчик: 1","0").replace("Подписчиков: "+str(Followers),"0")
        else:
            TypicalJobCost = soup.find("div", class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 dYJOPh ecpWHO gRCcss hui-cell").find_next_sibling().find_next_sibling().find_next_sibling().find_next_sibling().find("p", class_="sc-mwxddt-0 cZJFpr").text.replace("Подписчик: 1","0").replace("Подписчиков: "+str(Followers),"0")
    except:
        TypicalJobCost ="0"

    global ProjectsNumber
    ProjectsNumber= " 0"


        
    global Geographic
    a = "Нет географии"
    Geographic = a 
    try:
        
        Geographic = soup.find("p", class_= "sc-mwxddt-0 krnygu").find_next_sibling().find_next_sibling().text
    except:
        Geograpchic = "Нет географии"
    try:
        
        global Service
        Service = soup.find("p", class_= "sc-mwxddt-0 krnygu").text
    except:
        Service="Нет сервисов"
    try:
        CategoryFindAll = soup.find_all("button", class_="sc-1qppj2g-0 OaaNk hui-btn sc-1v049kn-0 gPleMa hz-track-me")
        global Category
        Category = [i.get_text() for i in CategoryFindAll]
    except:
        Category="Нет категорий"
    try:
        global FeedBack
        FeedBack = soup.find("span", class_="hz-star-rate__review-string").text.replace("1 отзыв","1")

    except:
        FeedBack="0"

        
    try:
        global Socials
        SocialsFindAll = soup.find("p","sc-mwxddt-0 jVkWgv").find_all("a",class_="sc-62xgu6-0 jtYTMy sc-mwxddt-0 dOJEup hui-link")
        Socials = [i.get("href") for i in SocialsFindAll]
        
    except:
        Socials ="Нет соцсетей"
    try:
        global Mark
        Mark = soup.find("span",class_="hz-star-rate__rating-number").text
    except:
        Mark="0.0"

    #print(str(URLneed)+" "+str(Address)+" "+str(BusinessName)+" "+str(PhoneNumber)+" "+str(Website)+" "+str(Contacts)+" "+str (Geographic)+" "+str(TypicalJobCost)+" "+str(ProjectsNumber).replace("Проектов: ","")+" "+str(Service)+" "+str(Category)+" "+str(FeedBack).replace("Отзывов: ", "")+" "+str(Followers)+" "+str(Socials)+" "+str(Mark))
    try:
        
        return str(URLneed),str(Address),str(BusinessName),str(PhoneNumber),str(Website),str(Contacts).replace("Контактное лицо: ",""),str(Geographic),str(TypicalJobCost),str(Service),str(Category),str(ProjectsNumber).replace("Проектов: ",""),str(FeedBack).replace("Отзывов: ", ""),str(Followers),str(Socials),str(Mark)
    except NameError:
        ProjectsNumber="0"
        return str(URLneed),str(Address),str(BusinessName),str(PhoneNumber),str(Website),str(Contacts).replace("Контактное лицо: ",""),str(Geographic),str(TypicalJobCost),str(Service),str(Category),str(ProjectsNumber).replace("Проектов: ",""),str(FeedBack).replace("Отзывов: ", ""),str(Followers),str(Socials),str(Mark)

fn = 'data3000.xlsx'
wb = load_workbook(fn)
ws = wb['data']

ws['A1'] = "Ссылка на карточку компании/персоны"
ws['B1'] = "Адрес"
ws['C1'] = "Название"
ws['D1'] = "Телефон"
ws['E1'] = "Сайт"
ws['F1'] = "Контактные лица"
ws['G1'] = "География работ"
ws['H1'] = "Цена"
ws['I1'] = "Услуги"
ws['J1'] = " Категории"
ws['K1'] = "Кол-во проектов"
ws['L1'] = " Кол-во отзывов"
ws['M1'] = " Кол-во подписчиков"
ws['N1'] = " Ссылки на соцсети"
ws['O1'] = " Рейтинг"

ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['C1'].font = Font(bold=True)
ws['D1'].font = Font(bold=True)
ws['E1'].font = Font(bold=True)
ws['F1'].font = Font(bold=True)
ws['G1'].font = Font(bold=True)
ws['H1'].font = Font(bold=True)
ws['I1'].font = Font(bold=True)
ws['J1'].font = Font(bold=True)
ws['K1'].font = Font(bold=True)
ws['L1'].font = Font(bold=True)
ws['M1'].font = Font(bold=True)
ws['N1'].font = Font(bold=True)
ws['O1'] .font = Font(bold=True)


def check_all_pages():
    URLD1  = "https://www.houzz.ru/professionals/dizayn-interyera/moskva-48-ru-probr0-bo~t_14028~r_524901" #Дизайнеры интерьера
    URLD2  = "https://www.houzz.ru/professionals/arkhitektory/moskva-48-ru-probr0-bo~t_14027~r_524901"#Архитекторы
    URLD3 = "https://www.houzz.ru/professionals/dekoratory-i-stilisty-interiera/moskva-48-ru-probr0-bo~t_30532~r_524901"#Декораторы и стилисты интерьера
    URLD4 = "https://www.houzz.ru/professionals/remont-i-otdelka-kvartir-i-domov/moskva-48-ru-probr0-bo~t_14083~r_524901"#Ремонт и отделка квартир и домов
    
    StringForAddress ="?fi="
    n=1
    
    finalString3 = URLD1

    response2 = requests.get(finalString3)
    soup2 = BeautifulSoup(response2.text, "lxml")
    
    
    NumberExperts = soup2.find("div", class_="hz-pro-search-controls__pagination mlm").find("span", class_="text-bold").next_element.next_element.next_element.next_element.next_element.next_element.text

    #print (NumberExperts)
    #NE = int(NumberExperts)
          
    Counter = 2

    #print(str(NE))

    for k in range(12421):
        
        finalString3 =str(URLD1)+"?fi="+str(k)
        response3 = requests.get(finalString3)
        soup3 = BeautifulSoup(response3.text, "lxml")
        try:
            AllLinks = soup3.find("li",class_="hz-pro-search-results__item").find_next_sibling().find_next_sibling().find_next_sibling().find_next_sibling().find_next_sibling().find_next_sibling().find_next_sibling().find("a",class_="hz-pro-ctl").get("href")
            print(str(AllLinks))
        except:
            print("ОШИБКА ССЫЛКИ")
        #LinksForProgramm = [i.get("href") for i in AllLinks]
        o=str(Counter)

        ws['A'+o],ws['B'+o],ws['C'+o],ws['D'+o],ws['E'+o],ws['F'+o],ws['G'+o],ws['H'+o],ws['I'+o],ws['J'+o] ,ws['K'+o],ws['L'+o],ws['M'+o],ws['N'+o],ws['O'+o]=ParsingPage(str(AllLinks))
        wb.save(fn)
 

            
        print(str(Counter))
        print("Время работы программы: "+str((time.time() - t_start)))
        Counter +=1
        
    #print(str(finalString3))
        
    return 0
check_all_pages()


wb.close()

print("Время работы программы: "+str((time.time() - t_start)))
